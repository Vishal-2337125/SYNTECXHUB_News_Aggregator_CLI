import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

class DataExporter:
    @staticmethod
    def export_clickable_excel(df, filename):
        path = f"{filename}"
        
        # Create Excel Hyperlink formula: =HYPERLINK("URL", "Headline")
        df['Clickable_Headline'] = df.apply(
            lambda row: f'=HYPERLINK("{row["url"]}", "{row["headline"].replace('"', "'")}")', 
            axis=1
        )
        
        # Save only the interactive column
        df[['Clickable_Headline']].to_excel(path, index=False)
        
        # Apply standard link styling (Blue & Underlined)
        wb = load_workbook(path)
        ws = wb.active
        link_font = Font(color="0563C1", underline="single")
        
        for cell in ws["A"][1:]:  # Skip header row
            cell.font = link_font
            
        wb.save(path)
        return path